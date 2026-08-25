from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
)
from openpyxl.utils import (
    get_column_letter,
)


def _value(
    value: Any,
) -> Any:

    if value is None:
        return ""

    if isinstance(
        value,
        (
            dict,
            list,
            tuple,
        ),
    ):

        return json.dumps(
            value,
            ensure_ascii=False,
            default=str,
        )

    return value


def _format_sheet(
    worksheet,
) -> None:

    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column_cells in (
        worksheet.columns
    ):

        max_length = 0

        column_index = (
            column_cells[
                0
            ].column
        )

        for cell in column_cells:

            try:
                length = len(
                    str(
                        cell.value
                        if cell.value
                        is not None
                        else ""
                    )
                )

                max_length = max(
                    max_length,
                    length,
                )

            except Exception:
                pass

        worksheet.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = min(
            max(
                max_length + 2,
                12,
            ),
            55,
        )


def _flatten(
    value: Any,
    prefix: str = "",
) -> list[
    tuple[
        str,
        Any,
    ]
]:

    output = []

    if isinstance(
        value,
        dict,
    ):

        for key, item in (
            value.items()
        ):

            new_prefix = (
                f"{prefix}.{key}"
                if prefix
                else str(
                    key
                )
            )

            output.extend(
                _flatten(
                    item,
                    new_prefix,
                )
            )

    elif isinstance(
        value,
        list,
    ):

        if not value:

            output.append(
                (
                    prefix,
                    "",
                )
            )

        else:

            for index, item in enumerate(
                value,
            ):

                new_prefix = (
                    f"{prefix}[{index}]"
                )

                output.extend(
                    _flatten(
                        item,
                        new_prefix,
                    )
                )

    else:

        output.append(
            (
                prefix,
                value,
            )
        )

    return output


def generate_excel_report(
    report_data: dict[str, Any],
    output_path: str,
) -> str:

    Path(
        output_path
    ).parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    # ========================================================
    # Report Summary
    # ========================================================

    summary_sheet = workbook.active

    summary_sheet.title = (
        "Report Summary"
    )

    summary_sheet.append(
        [
            "Field",
            "Value",
        ]
    )

    summary_sheet.append(
        [
            "Title",
            report_data.get(
                "title",
                "",
            ),
        ]
    )

    summary_sheet.append(
        [
            "Dataset ID",
            report_data.get(
                "dataset_id",
                "",
            ),
        ]
    )

    summary_sheet.append(
        [
            "Generated At",
            report_data.get(
                "generated_at",
                "",
            ),
        ]
    )

    summary = report_data.get(
        "summary",
        {},
    )

    summary_sheet.append(
        [
            "Analysis Count",
            summary.get(
                "analysis_count",
                0,
            ),
        ]
    )

    summary_sheet.append(
        [
            "Visualization Count",
            summary.get(
                "visualization_count",
                0,
            ),
        ]
    )

    summary_sheet.append(
        [
            (
                "Smart Interpretation "
                "Available"
            ),
            summary.get(
                (
                    "smart_"
                    "interpretation_"
                    "available"
                ),
                False,
            ),
        ]
    )

    _format_sheet(
        summary_sheet
    )

    # ========================================================
    # Dataset Information
    # ========================================================

    dataset_sheet = (
        workbook.create_sheet(
            "Dataset Information"
        )
    )

    dataset_sheet.append(
        [
            "Field",
            "Value",
        ]
    )

    for key, value in _flatten(
        report_data.get(
            "dataset",
            {},
        )
    ):

        dataset_sheet.append(
            [
                key,
                _value(
                    value
                ),
            ]
        )

    _format_sheet(
        dataset_sheet
    )

    # ========================================================
    # Statistical Analyses
    # ========================================================

    analysis_sheet = (
        workbook.create_sheet(
            "Analyses"
        )
    )

    analysis_sheet.append(
        [
            "Analysis #",
            "Field",
            "Value",
        ]
    )

    analyses = report_data.get(
        "analyses",
        [],
    )

    for index, analysis in enumerate(
        analyses,
        1,
    ):

        for field, value in _flatten(
            analysis
        ):

            analysis_sheet.append(
                [
                    index,
                    field,
                    _value(
                        value
                    ),
                ]
            )

    _format_sheet(
        analysis_sheet
    )

    # ========================================================
    # Visualizations
    # ========================================================

    visualization_sheet = (
        workbook.create_sheet(
            "Visualizations"
        )
    )

    visualization_sheet.append(
        [
            "Visualization #",
            "Field",
            "Value",
        ]
    )

    visualizations = (
        report_data.get(
            "visualizations",
            [],
        )
    )

    for index, visualization in (
        enumerate(
            visualizations,
            1,
        )
    ):

        for field, value in _flatten(
            visualization
        ):

            visualization_sheet.append(
                [
                    index,
                    field,
                    _value(
                        value
                    ),
                ]
            )

    _format_sheet(
        visualization_sheet
    )

    # ========================================================
    # Smart Interpretation
    # ========================================================

    interpretation_sheet = (
        workbook.create_sheet(
            "Smart Interpretation"
        )
    )

    interpretation_sheet.append(
        [
            "Field",
            "Value",
        ]
    )

    smart = report_data.get(
        "smart_interpretation"
    )

    if smart:

        for field, value in _flatten(
            smart
        ):

            interpretation_sheet.append(
                [
                    field,
                    _value(
                        value
                    ),
                ]
            )

    else:

        interpretation_sheet.append(
            [
                "status",
                "No smart interpretation "
                "available.",
            ]
        )

    _format_sheet(
        interpretation_sheet
    )

    # ========================================================
    # Warnings
    # ========================================================

    warning_sheet = (
        workbook.create_sheet(
            "Warnings"
        )
    )

    warning_sheet.append(
        [
            "Severity",
            "Warning",
        ]
    )

    warnings = []

    if (
        smart
        and smart.get(
            "available"
        )
    ):

        interpretation_response = (
            smart.get(
                "interpretation",
                {},
            )
        )

        interpretation = (
            interpretation_response.get(
                "interpretation",
                {},
            )
        )

        warnings = interpretation.get(
            "cautions",
            [],
        )

    for warning in warnings:

        warning_sheet.append(
            [
                "Caution",
                str(
                    warning
                ),
            ]
        )

    if not warnings:

        warning_sheet.append(
            [
                "Info",
                (
                    "No statistical warnings "
                    "were returned."
                ),
            ]
        )

    _format_sheet(
        warning_sheet
    )

    workbook.save(
        output_path
    )

    return output_path
